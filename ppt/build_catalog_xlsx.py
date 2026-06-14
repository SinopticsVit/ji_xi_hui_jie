#!/usr/bin/env python3
"""Build Excel catalog from catalog-ru.md with embedded images."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

HERE = Path(__file__).resolve().parent
MATERIALS_DIR = HERE / "maretials"
CATALOG_MD = MATERIALS_DIR / "catalog-ru.md"
TEMPLATE_XLSX = MATERIALS_DIR / "catalog-template.xlsx"
FILLED_XLSX = MATERIALS_DIR / "catalog-ru.xlsx"

HEADERS = [
    "ID",
    "Раздел",
    "Блок",
    "Подблок",
    "Группа / серия",
    "№ в группе",
    "Артикул",
    "Наименование / тип",
    "Примечание",
    "Тип записи",
    "Фото",
    "Путь к файлу",
]

PHOTO_COL = 11  # K
ROW_HEIGHT = 68
IMAGE_MAX_W = 110
IMAGE_MAX_H = 80


@dataclass
class CatalogRow:
    section: str = ""
    block: str = ""
    subblock: str = ""
    group: str = ""
    num: str = ""
    sku: str = ""
    name: str = ""
    note: str = ""
    record_type: str = "group"
    image_path: str | None = None
    merge_key: str | None = None


def extract_image_path(line: str) -> str | None:
    m = re.search(r"!\[[^\]]*\]\(([^)]+)\)", line)
    return m.group(1).strip() if m else None


def extract_thumb_path(cell: str) -> str | None:
    return extract_image_path(cell)


def parse_catalog_md(text: str) -> list[CatalogRow]:
    lines = text.splitlines()
    rows: list[CatalogRow] = []

    section = ""
    block = ""
    subblock = ""
    subblock_title = ""
    section_image: str | None = None
    current_group = ""
    pending_bullets: list[str] = []
    in_product_table = False

    def flush_bullets() -> None:
        nonlocal pending_bullets
        if not pending_bullets or not subblock:
            pending_bullets = []
            return
        for bullet in pending_bullets:
            rows.append(
                CatalogRow(
                    section=section,
                    block=block,
                    subblock=subblock,
                    group=current_group or subblock_title,
                    sku="—",
                    name=bullet,
                    record_type="group_item",
                    image_path=section_image,
                    merge_key=f"{subblock}:{section_image}",
                )
            )
        pending_bullets = []

    def add_group_row(name: str, note: str = "") -> None:
        rows.append(
            CatalogRow(
                section=section,
                block=block,
                subblock=subblock,
                group=current_group or subblock_title,
                sku="—",
                name=name,
                note=note,
                record_type="group",
                image_path=section_image,
                merge_key=f"{subblock}:{section_image}",
            )
        )

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if stripped.startswith("## Раздел A"):
            flush_bullets()
            section = "A"
            i += 1
            continue
        if stripped.startswith("## Раздел B"):
            flush_bullets()
            section = "B"
            i += 1
            continue

        if stripped.startswith("### "):
            flush_bullets()
            m = re.match(r"### (A\.\d+|B\.\d+)\s", stripped)
            if m:
                block = m.group(1)
            i += 1
            continue

        if stripped.startswith("#### "):
            flush_bullets()
            in_product_table = False
            title = stripped[5:].strip()
            subblock_m = re.match(r"((?:A|B)\.\d+\.\d+)", title)
            subblock = subblock_m.group(1) if subblock_m else title.split()[0]
            subblock_title = title
            current_group = ""
            section_image = None
            i += 1
            while i < len(lines):
                nxt = lines[i].strip()
                if nxt.startswith("#"):
                    break
                img = extract_image_path(nxt)
                if img and section_image is None:
                    section_image = img
                if nxt.startswith("|") and "Артикул" in nxt:
                    break
                if nxt.startswith("|") and "Тип" in nxt and "Артикул" not in nxt:
                    break
                if nxt.startswith("- "):
                    pending_bullets.append(nxt[2:].strip())
                    i += 1
                    continue
                if nxt and not nxt.startswith("!") and not nxt.startswith("*") and not nxt.startswith("**"):
                    if not nxt.startswith("|") and "Итого" not in nxt:
                        add_group_row(subblock_title, nxt)
                        section_image = section_image  # keep for potential items
                i += 1
            continue

        if stripped.startswith("##### "):
            flush_bullets()
            current_group = stripped[6:].strip()
            in_product_table = False
            i += 1
            continue

        if stripped.startswith("| № |") and "Артикул" in stripped:
            in_product_table = True
            i += 2
            while i < len(lines):
                row_line = lines[i].strip()
                if not row_line.startswith("|"):
                    in_product_table = False
                    break
                if row_line.startswith("| № |"):
                    i += 1
                    continue
                parts = [p.strip() for p in row_line.strip("|").split("|")]
                if len(parts) >= 5 and parts[1] != "Артикул":
                    thumb = extract_thumb_path(parts[2]) if len(parts) > 2 else None
                    rows.append(
                        CatalogRow(
                            section=section,
                            block=block,
                            subblock=subblock,
                            group=current_group,
                            num=parts[0],
                            sku=parts[1],
                            name=parts[3],
                            note=parts[4] if parts[4] != "—" else "",
                            record_type="product",
                            image_path=thumb or section_image,
                            merge_key=None,
                        )
                    )
                i += 1
            continue

        if stripped.startswith("| № |") and "Тип" in stripped and "Артикул" not in stripped:
            i += 2
            item_rows: list[CatalogRow] = []
            while i < len(lines):
                row_line = lines[i].strip()
                if not row_line.startswith("|"):
                    break
                parts = [p.strip() for p in row_line.strip("|").split("|")]
                if len(parts) >= 2 and parts[0].isdigit():
                    item_rows.append(
                        CatalogRow(
                            section=section,
                            block=block,
                            subblock=subblock,
                            group=current_group or subblock_title,
                            num=parts[0],
                            sku="—",
                            name=parts[1],
                            record_type="group_item",
                            image_path=section_image,
                            merge_key=f"{subblock}:{section_image}",
                        )
                    )
                i += 1
            rows.extend(item_rows)
            continue

        if stripped.startswith("- ") and subblock and not in_product_table:
            pending_bullets.append(stripped[2:].strip())
        i += 1

    flush_bullets()
    return rows


def style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    widths = {
        "A": 10,
        "B": 8,
        "C": 10,
        "D": 12,
        "E": 28,
        "F": 10,
        "G": 18,
        "H": 36,
        "I": 24,
        "J": 12,
        "K": 18,
        "L": 42,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L1"


def add_help_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Справка")
    lines = [
        ("Описание колонок", ""),
        ("ID", "Сквозной идентификатор CAT-NNN"),
        ("Раздел", "A — запирающие системы; B — крепёж"),
        ("Блок / Подблок", "Иерархия из catalog-ru.md (A.2.1 …)"),
        ("Группа / серия", "Серия HB2, TR-K101 или название секции"),
        ("Артикул", "SKU; «—» для групповых позиций"),
        ("Тип записи", "product — позиция с артикулом; group — секция; group_item — элемент группы"),
        ("Фото", "Встроенное изображение; для group/group_item — объединённая ячейка на группу"),
        ("", ""),
        ("Пересборка", "python build_catalog_xlsx.py"),
        ("Источник", "maretials/catalog-ru.md"),
    ]
    for r, (a, b) in enumerate(lines, start=1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if r == 1:
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def scale_image(path: Path) -> tuple[Path, int, int]:
    with PILImage.open(path) as im:
        w, h = im.size
        scale = min(IMAGE_MAX_W / w, IMAGE_MAX_H / h, 1.0)
        nw, nh = max(1, int(w * scale)), max(1, int(h * scale))
        if (nw, nh) != (w, h):
            im = im.resize((nw, nh), PILImage.Resampling.LANCZOS)
            tmp = path.parent / f"_xlsx_{path.name}"
            im.save(tmp, format="PNG")
            return tmp, nw, nh
        return path, w, h


def embed_image(ws, row: int, path: Path) -> None:
    if not path.is_file():
        return
    img_path, w, h = scale_image(path)
    xl_img = XLImage(str(img_path))
    xl_img.width = w
    xl_img.height = h
    anchor = f"{get_column_letter(PHOTO_COL)}{row}"
    ws.add_image(xl_img, anchor)
    ws.row_dimensions[row].height = max(ROW_HEIGHT, h * 0.75)


def write_catalog_sheet(ws, catalog_rows: list[CatalogRow], with_data: bool) -> None:
    style_header(ws)
    if not with_data:
        return

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    merge_ranges: dict[str, list[int]] = {}

    for idx, row in enumerate(catalog_rows, start=1):
        excel_row = idx + 1
        ws.cell(row=excel_row, column=1, value=f"CAT-{idx:03d}")
        ws.cell(row=excel_row, column=2, value=row.section)
        ws.cell(row=excel_row, column=3, value=row.block)
        ws.cell(row=excel_row, column=4, value=row.subblock)
        ws.cell(row=excel_row, column=5, value=row.group)
        ws.cell(row=excel_row, column=6, value=row.num)
        ws.cell(row=excel_row, column=7, value=row.sku)
        ws.cell(row=excel_row, column=8, value=row.name)
        ws.cell(row=excel_row, column=9, value=row.note)
        ws.cell(row=excel_row, column=10, value=row.record_type)
        ws.cell(row=excel_row, column=12, value=row.image_path or "")

        for col in range(1, 13):
            c = ws.cell(row=excel_row, column=col)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)

        if row.image_path:
            full = MATERIALS_DIR / row.image_path.replace("/", "\\")
            if not full.is_file():
                full = MATERIALS_DIR / row.image_path
            if row.record_type == "product":
                embed_image(ws, excel_row, full)
            elif row.merge_key:
                merge_ranges.setdefault(row.merge_key, []).append(excel_row)
            else:
                embed_image(ws, excel_row, full)

    for key, excel_rows in merge_ranges.items():
        if not excel_rows:
            continue
        r1, r2 = min(excel_rows), max(excel_rows)
        if r2 > r1:
            ws.merge_cells(
                start_row=r1,
                start_column=PHOTO_COL,
                end_row=r2,
                end_column=PHOTO_COL,
            )
        first_row = catalog_rows[r1 - 2]
        if first_row.image_path:
            full = MATERIALS_DIR / first_row.image_path
            embed_image(ws, r1, full)


def build_workbook(catalog_rows: list[CatalogRow], with_data: bool) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"
    write_catalog_sheet(ws, catalog_rows, with_data=with_data)
    add_help_sheet(wb)
    return wb


def verify_rows(catalog_rows: list[CatalogRow]) -> None:
    products = [r for r in catalog_rows if r.record_type == "product"]
    skus = {r.sku for r in products}
    if len(skus) != 82:
        raise ValueError(f"Expected 82 product SKUs, got {len(skus)}")
    missing = []
    for r in catalog_rows:
        if r.image_path:
            p = MATERIALS_DIR / r.image_path
            if not p.is_file():
                missing.append(r.image_path)
    if missing:
        raise FileNotFoundError(f"Missing images: {missing[:5]} ... ({len(missing)} total)")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--template-only", action="store_true")
    ap.add_argument("--filled-only", action="store_true")
    args = ap.parse_args()

    text = CATALOG_MD.read_text(encoding="utf-8")
    catalog_rows = parse_catalog_md(text)
    verify_rows(catalog_rows)
    print(f"[info] parsed {len(catalog_rows)} rows ({sum(1 for r in catalog_rows if r.record_type=='product')} products)")

    if not args.filled_only:
        wb = build_workbook([], with_data=False)
        wb.save(TEMPLATE_XLSX)
        print(f"[done] template: {TEMPLATE_XLSX}")

    if not args.template_only:
        wb = build_workbook(catalog_rows, with_data=True)
        wb.save(FILLED_XLSX)
        print(f"[done] catalog: {FILLED_XLSX}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
