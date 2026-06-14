#!/usr/bin/env python3
"""Build Excel catalog from zhengling-machinery/catalog-ru.md."""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

HERE = Path(__file__).resolve().parent
MATERIALS_DIR = HERE / "maretials"
ZHENGLING_DIR = MATERIALS_DIR / "zhengling-machinery"
CATALOG_MD = ZHENGLING_DIR / "catalog-ru.md"
TEMPLATE_XLSX = ZHENGLING_DIR / "catalog-template.xlsx"
FILLED_XLSX = ZHENGLING_DIR / "catalog-ru.xlsx"

HEADERS = [
    "ID", "Раздел", "Блок", "Подблок", "Группа / серия", "№ в группе",
    "Артикул", "Наименование / тип", "Примечание", "Тип записи", "Фото", "Путь к файлу",
]
PHOTO_COL = 11
ROW_HEIGHT = 68
IMAGE_MAX_W = 110
IMAGE_MAX_H = 80


@dataclass
class CatalogRow:
    section: str = ""
    block: str = ""
    subblock: str = ""
    group: str = ""
    num: str = ""
    sku: str = ""
    name: str = ""
    note: str = ""
    record_type: str = "group"
    image_path: str | None = None
    merge_key: str | None = None


def extract_image_path(line: str) -> str | None:
    m = re.search(r"!\[[^\]]*\]\(([^)]+)\)", line)
    return m.group(1).strip() if m else None


def parse_catalog_md(text: str) -> list[CatalogRow]:
    lines = text.splitlines()
    rows: list[CatalogRow] = []
    section = ""
    block = ""
    subblock = ""
    subblock_title = ""
    section_image: str | None = None
    pending_bullets: list[str] = []
    pending_note: list[str] = []

    def flush_section() -> None:
        nonlocal pending_bullets, pending_note, section_image
        if not subblock:
            pending_bullets = []
            pending_note = []
            return
        note = "\n".join(pending_note).strip()
        if pending_bullets:
            for i, bullet in enumerate(pending_bullets, start=1):
                rows.append(
                    CatalogRow(
                        section=section,
                        block=block,
                        subblock=subblock,
                        group=subblock_title,
                        num=str(i),
                        sku="—",
                        name=bullet,
                        note=note if i == 1 else "",
                        record_type="group_item",
                        image_path=section_image,
                        merge_key=f"{subblock}:{section_image}",
                    )
                )
        else:
            rows.append(
                CatalogRow(
                    section=section,
                    block=block,
                    subblock=subblock,
                    group=subblock_title,
                    sku="—",
                    name=subblock_title,
                    note=note,
                    record_type="group",
                    image_path=section_image,
                    merge_key=f"{subblock}:{section_image}",
                )
            )
        pending_bullets = []
        pending_note = []
        section_image = None

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if stripped.startswith("## Раздел"):
            flush_section()
            section = "C"
            i += 1
            continue

        if stripped.startswith("### "):
            flush_section()
            m = re.match(r"### (C\.\d+)\s", stripped)
            if m:
                block = m.group(1)
            i += 1
            continue

        if stripped.startswith("#### "):
            flush_section()
            title = stripped[5:].strip()
            subblock_m = re.match(r"(C\.\d+\.\d+)", title)
            subblock = subblock_m.group(1) if subblock_m else title.split()[0]
            subblock_title = re.sub(r"^C\.\d+\.\d+\s*", "", title)
            section_image = None
            pending_note = []
            pending_bullets = []
            i += 1
            while i < len(lines):
                nxt = lines[i].strip()
                if nxt.startswith("#"):
                    break
                img = extract_image_path(nxt)
                if img and section_image is None:
                    section_image = img
                elif nxt.startswith("- "):
                    pending_bullets.append(nxt[2:].strip())
                elif nxt.startswith("|") and not nxt.startswith("| ID |"):
                    pending_note.append(nxt)
                elif nxt and not nxt.startswith("!") and not nxt.startswith("---"):
                    pending_note.append(nxt)
                i += 1
            flush_section()
            continue

        i += 1

    flush_section()
    return rows


def style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    widths = {"A": 10, "B": 8, "C": 10, "D": 12, "E": 28, "F": 10,
              "G": 18, "H": 36, "I": 24, "J": 12, "K": 18, "L": 42}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:L1"


def add_help_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Справка")
    lines = [
        ("Описание колонок", ""),
        ("ID", "Сквозной идентификатор CAT-NNN"),
        ("Раздел", "C — компоненты рулевой системы"),
        ("Тип записи", "group — секция; group_item — элемент списка"),
        ("Пересборка", "python build_zhengling_catalog_xlsx.py"),
        ("Источник", "zhengling-machinery/catalog-ru.md"),
    ]
    for r, (a, b) in enumerate(lines, start=1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if r == 1:
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def scale_image(path: Path) -> tuple[Path, int, int]:
    with PILImage.open(path) as im:
        w, h = im.size
        scale = min(IMAGE_MAX_W / w, IMAGE_MAX_H / h, 1.0)
        nw, nh = max(1, int(w * scale)), max(1, int(h * scale))
        if (nw, nh) != (w, h):
            im = im.resize((nw, nh), PILImage.Resampling.LANCZOS)
            tmp = path.parent / f"_xlsx_{path.name}"
            im.save(tmp, format="PNG")
            return tmp, nw, nh
        return path, w, h


def embed_image(ws, row: int, path: Path) -> None:
    if not path.is_file():
        return
    img_path, w, h = scale_image(path)
    xl_img = XLImage(str(img_path))
    xl_img.width = w
    xl_img.height = h
    ws.add_image(xl_img, f"{get_column_letter(PHOTO_COL)}{row}")
    ws.row_dimensions[row].height = max(ROW_HEIGHT, h * 0.75)


def write_catalog_sheet(ws, catalog_rows: list[CatalogRow], with_data: bool) -> None:
    style_header(ws)
    if not with_data:
        return
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    merge_ranges: dict[str, list[int]] = {}

    for idx, row in enumerate(catalog_rows, start=1):
        excel_row = idx + 1
        ws.cell(row=excel_row, column=1, value=f"CAT-{idx:03d}")
        for col, val in enumerate(
            [row.section, row.block, row.subblock, row.group, row.num,
             row.sku, row.name, row.note, row.record_type, "", row.image_path or ""],
            start=2,
        ):
            ws.cell(row=excel_row, column=col, value=val)
        for col in range(1, 13):
            c = ws.cell(row=excel_row, column=col)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
        if row.image_path:
            full = ZHENGLING_DIR / row.image_path
            if row.merge_key:
                merge_ranges.setdefault(row.merge_key, []).append(excel_row)
            else:
                embed_image(ws, excel_row, full)

    for key, excel_rows in merge_ranges.items():
        if not excel_rows:
            continue
        r1, r2 = min(excel_rows), max(excel_rows)
        if r2 > r1:
            ws.merge_cells(start_row=r1, start_column=PHOTO_COL, end_row=r2, end_column=PHOTO_COL)
        first = catalog_rows[r1 - 2]
        if first.image_path:
            embed_image(ws, r1, ZHENGLING_DIR / first.image_path)


def verify_rows(catalog_rows: list[CatalogRow]) -> None:
    if len(catalog_rows) < 35:
        raise ValueError(f"Expected at least 35 rows, got {len(catalog_rows)}")
    slides = set()
    missing = []
    for r in catalog_rows:
        if r.image_path:
            slides.add(r.image_path)
            p = ZHENGLING_DIR / r.image_path
            if not p.is_file():
                missing.append(r.image_path)
    if len(slides) != 35:
        raise ValueError(f"Expected 35 unique slide images, got {len(slides)}")
    if missing:
        raise FileNotFoundError(f"Missing: {missing[:5]}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--template-only", action="store_true")
    ap.add_argument("--filled-only", action="store_true")
    args = ap.parse_args()

    text = CATALOG_MD.read_text(encoding="utf-8")
    catalog_rows = parse_catalog_md(text)
    verify_rows(catalog_rows)
    n_imgs = len({r.image_path for r in catalog_rows if r.image_path})
    print(f"[info] parsed {len(catalog_rows)} rows, {n_imgs} images")

    if not args.filled_only:
        wb = Workbook()
        ws = wb.active
        ws.title = "Каталог"
        write_catalog_sheet(ws, [], with_data=False)
        add_help_sheet(wb)
        wb.save(TEMPLATE_XLSX)
        print(f"[done] template: {TEMPLATE_XLSX}")

    if not args.template_only:
        wb = Workbook()
        ws = wb.active
        ws.title = "Каталог"
        write_catalog_sheet(ws, catalog_rows, with_data=True)
        add_help_sheet(wb)
        wb.save(FILLED_XLSX)
        print(f"[done] catalog: {FILLED_XLSX}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
